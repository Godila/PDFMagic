"""
excel_handler.py — чтение параметров из Excel-шаблона и запись результатов.

Ожидаемая структура входного .xlsx:
  Лист "Список параметров":
    Строка 1  — заголовки (A=№, B=Приписка, C=Тип, D=Параметр,
                            E=Комментарий, F=Ключевые слова,
                            G=Единицы измерения, H=Наименования)
    Строки 2+ — данные параметров

  Лист "параметры по концепциям":
    Строка 1  — заголовки (A=№, C=Параметр, D+ = данные по концепциям)
    Строки 2+ — значения параметров
"""
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Константы имён листов (с учётом регистра из файла)
# ---------------------------------------------------------------------------
SHEET_PARAMS_LIST = "Список параметров"
SHEET_CONCEPTS = "параметры по концепциям"


@dataclass
class Parameter:
    """Один параметр из листа «Список параметров»."""
    row_index: int          # Порядковый номер строки (1-based, без заголовка)
    number: str             # Столбец A — №
    attribution: str        # Столбец B — Приписка параметра
    param_type: str         # Столбец C — Тип параметра
    name: str               # Столбец D — Параметр (основное название)
    comment: str            # Столбец E — Комментарий
    keywords: str           # Столбец F — Ключевые слова
    units: str              # Столбец G — Единицы измерения
    column_names: str       # Столбец H — Наименования колонок/строк


def load_parameters(xlsx_path: str) -> list[Parameter]:
    """
    Читает список параметров из листа «Список параметров».
    Возвращает список объектов Parameter в порядке строк.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Найти лист (нечувствительно к регистру)
    sheet = _find_sheet(wb, SHEET_PARAMS_LIST)

    params: list[Parameter] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        # Пропустить полностью пустые строки
        if all(cell is None or str(cell).strip() == "" for cell in row[:6]):
            continue

        def _val(cell) -> str:
            return str(cell).strip() if cell is not None else ""

        params.append(
            Parameter(
                row_index=row_idx,
                number=_val(row[0]),     # A
                attribution=_val(row[1]),# B
                param_type=_val(row[2]), # C
                name=_val(row[3]),       # D
                comment=_val(row[4]),    # E
                keywords=_val(row[5]),   # F
                units=_val(row[6]) if len(row) > 6 else "",  # G
                column_names=_val(row[7]) if len(row) > 7 else "",  # H
            )
        )

    wb.close()
    print(f"[excel_handler] Загружено параметров: {len(params)}")
    return params


def write_results(
    xlsx_path: str,
    parameters: list[Parameter],
    extracted_values: dict[int, str | None],
    column_header: str,
    output_path: str | None = None,
) -> str:
    """
    Добавляет новый столбец с результатами в лист «параметры по концепциям».

    Args:
        xlsx_path:         Путь к исходному .xlsx файлу.
        parameters:        Список параметров (из load_parameters).
        extracted_values:  Словарь {row_index: значение_или_None}.
        column_header:     Заголовок нового столбца (имя PDF без расширения).
        output_path:       Куда сохранить результат (по умолчанию — рядом с оригиналом).

    Returns:
        Путь к сохранённому файлу.
    """
    if output_path is None:
        stem = Path(xlsx_path).stem
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path(xlsx_path).parent / f"{stem}_output_{date_str}.xlsx")

    # Копируем оригинал, чтобы не изменять его
    shutil.copy2(xlsx_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    sheet = _find_sheet(wb, SHEET_CONCEPTS)

    # Найти первый пустой столбец (после последней заполненной строки заголовка)
    new_col = _find_next_empty_column(sheet, header_row=1)
    col_letter = get_column_letter(new_col)

    print(f"[excel_handler] Записываем в столбец {col_letter} ({column_header})")

    # Заголовок
    header_cell = sheet[f"{col_letter}1"]
    header_cell.value = column_header
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(wrap_text=True, vertical="top")
    header_cell.fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )

    # Строим маппинг: число из столбца A -> номер строки в листе концепций
    # (порядок строк в "параметры по концепциям" ОТЛИЧАЕТСЯ от "Список параметров")
    num_to_sheet_row: dict[str, int] = {}
    for r in range(2, sheet.max_row + 2):
        cell_num = sheet.cell(r, 1).value
        if cell_num is not None:
            num_to_sheet_row[str(cell_num).strip()] = r

    written = 0
    not_found = []
    for param in parameters:
        value = extracted_values.get(param.row_index)
        if value is None:
            value = ""

        sheet_row = num_to_sheet_row.get(str(param.number).strip())
        if sheet_row is None:
            not_found.append(param.number)
            continue

        cell = sheet[f"{col_letter}{sheet_row}"]
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        written += 1

    if not_found:
        print(f"[excel_handler] WARN: не найдены строки для №: {not_found}")

    # Установить ширину нового столбца
    sheet.column_dimensions[col_letter].width = 20

    wb.save(output_path)
    wb.close()
    print(f"[excel_handler] Сохранено {written} значений -> {output_path}")
    return output_path


def _find_sheet(wb: openpyxl.Workbook, name: str):
    """Ищет лист по имени (нечувствительно к регистру)."""
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() == name.lower().strip():
            return wb[sheet_name]
    # Если точного совпадения нет — ищем частичное
    for sheet_name in wb.sheetnames:
        if name.lower().replace(" ", "") in sheet_name.lower().replace(" ", ""):
            return wb[sheet_name]
    available = ", ".join(f'"{s}"' for s in wb.sheetnames)
    raise ValueError(f'Лист "{name}" не найден. Доступные листы: {available}')


def _find_next_empty_column(sheet, header_row: int = 1) -> int:
    """
    Возвращает индекс (1-based) первого пустого столбца в строке заголовка.
    """
    max_col = sheet.max_column or 1
    for col in range(1, max_col + 2):
        cell = sheet.cell(row=header_row, column=col)
        if cell.value is None or str(cell.value).strip() == "":
            return col
    return max_col + 1


def parameters_to_prompt_text(parameters: list[Parameter]) -> str:
    """
    Формирует текстовое описание всех параметров для системного промпта LLM.
    """
    lines = []
    for p in parameters:
        parts = [f"#{p.row_index}"]
        if p.name:
            parts.append(f"Параметр: {p.name}")
        if p.attribution:
            parts.append(f"Категория: {p.attribution}")
        if p.param_type:
            parts.append(f"Тип: {p.param_type}")
        if p.keywords:
            parts.append(f"Ключевые слова: {p.keywords}")
        if p.units:
            parts.append(f"Единицы: {p.units}")
        if p.column_names:
            parts.append(f"Поле: {p.column_names}")
        if p.comment:
            parts.append(f"Примечание: {p.comment}")
        lines.append(" | ".join(parts))
    return "\n".join(lines)
